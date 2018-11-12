import h5py
import numpy as np
from openpyxl import load_workbook
from read_dataset import scan
#from pymvpa import mvpa2.mappers as mappers

# This method reads the data that I received from Jonas Kaplan.
# It is described in Dehghani et al. 2017
# Paper: https://onlinelibrary.wiley.com/doi/epdf/10.1002/hbm.23814


# It consists of fMRI data from 90 subjects (30 from three different native languages each: English, Mandarin, Farsi)
#  who read 40 short personal stories (145-155 words long).
# The data is already preprocessed, see section "fMRI Data Preprocessing" in the paper.
# Format: a matrix of 30 x 40 x 212018, which corresponds to subjects X stories X voxels.
# Note: the data for subject 30 is empty for English! Don't know why.
# Language should be english, chinese, or farsi
def read_all(data_dir, language = "english"):
    datafile = data_dir + "/30_" + language + "_storydata_masked.hd5"

    data = h5py.File(datafile, 'r')
    datamatrix = np.array(data["__unnamed__"][()])
    stimulifile = data_dir + '/StoryKey.xlsx'
    stimuli = load_workbook(stimulifile).active

    stories = []
    for story_id in range(2, stimuli.max_row + 1):
        # The first 7 columns contain irrelevant information
        context = stimuli.cell(row=story_id, column=8).value.strip()
        seg1 = stimuli.cell(row=story_id, column=9).value.strip()
        seg2 = stimuli.cell(row=story_id, column=10).value.strip()
        seg3 = stimuli.cell(row=story_id, column=11).value.strip()
        story = seg1 + " " + seg2 + " " + seg3
        story = story.replace("  ", " ")
        stories.append((context, story))

    scan_events = []
    for subject in range(0, datamatrix.shape[0]):
        for i in range(0, datamatrix.shape[1]):
            event = scan.ScanEvent()

            context = stories[i][0]
            text = stories[i][1]
            # For this dataset, the brain activation has already been averaged over the four different stimuli,
            # so I put the whole concatenated input into sentences.
            # I do not yet have a theory on whether it makes sense to include the context primer or not.
            event.sentences = context + ". " + text

            event.block = i
            event.subject_id = str(subject)
            event.scan = datamatrix[subject][i]
            scan_events.append(event)
    return scan_events

# def get_voxel_to_region_mapping(mapperfile, data):
    #TODO: I NEED PYMVPA FOR THIS
    #mapper = mvpa2.mappers.flatten.FlattenMapper(h5py.File(mapperfile, 'r'))
    #print(mapper)
    #print(mapper.reverse(data).shape())
    #return mapper.reverse(data)
